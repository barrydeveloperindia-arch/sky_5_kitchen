import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

# Creates a new workbook
wb = openpyxl.Workbook()

# --- Sheet 1: Master Menu Catalog ---
ws_menu = wb.active
ws_menu.title = "Master Menu"

# Headers
headers = ["ID", "Category", "Dish Name", "Description", "Sale Price (INR)", "MRP (INR)", "Is Best Seller", "Image"]
ws_menu.append(headers)

# Data (replicating data/combos.js)
menu_data = [
    [1, "Combos", "Dal Tadka Combo", "Dal Tadka + Jeera Rice + 2 Butter Roti.", 799, 899, "Yes", "dal_tadka_roti_rice_combo_1770013924380.png"],
    [2, "Combos", "Rajma Chawal Combo", "Home-style Rajma + Steamed Rice.", 649, 749, "No", "rajma_chawal_combo_1770013948564.png"],
    [4, "Combos", "Paneer Butter Masala Combo", "Paneer Butter Masala + Steamed Rice.", 849, 949, "Yes", "paneer_butter_masala_combo_bowl_1770012612956.png"],
    [11, "Soups", "Veg Soup", "A warm, comforting bowl of freshly prepared vegetable soup, made in our boutique hotel kitchen with seasonal vegetables and gentle spices. Perfect for a light yet satisfying meal.", 149, 199, "No", "Placeholder"],
    [12, "Soups", "Veg Sweet Corn Soup", "Classic hotel-style sweet corn soup with a rich, mildly sweet flavour and hearty texture. Comfort food that’s easy on the stomach and full of taste.", 169, 219, "No", "Placeholder"],
    [13, "Starters", "Tandoori Malai Broccoli", "Creamy tandoori-style broccoli.", 349, 399, "No", "paneer_tikka_skewer_1770013745197.png"],
    [14, "Starters", "Paneer Tikka", "Grilled paneer with spices.", 329, 389, "Yes", "paneer_tikka_skewer_1770013745197.png"],
    [15, "Starters", "Chilli Paneer", "Paneer tossed in spicy sauce.", 349, 399, "Yes", "chilli_paneer_fried_rice_combo_1770012631322.png"],
    [16, "Starters", "Veg Spring Roll", "Crispy vegetable rolls.", 249, 299, "No", "hakka_noodles_veg_1770013764042.png"],
    [17, "Rice & Noodles", "Veg Fried Rice", "Long-grain rice tossed with fresh vegetables and light seasoning, cooked in hotel-style for a balanced, non-greasy flavour.", 249, 299, "No", "chilli_paneer_fried_rice_combo_1770012631322.png"],
    [18, "Rice & Noodles", "Veg Hakka Noodles", "Soft noodles stir-fried with fresh vegetables, soy and subtle spices. A comforting Indo-Chinese favourite prepared fresh to order.", 249, 299, "Yes", "hakka_noodles_veg_1770013764042.png"],
    [19, "Rice & Noodles", "Veg Chilli Garlic Noodles", "Flavourful noodles tossed with vegetables, garlic and chilli for a mildly spicy kick, prepared in classic hotel-style.", 269, 319, "No", "hakka_noodles_veg_1770013764042.png"],
    [26, "Rice & Noodles", "Vegetable Biryani", "Fragrant basmati rice cooked with fresh vegetables and mild spices, served with cooling raita. Light, aromatic and satisfying.", 349, 399, "No", "dal_tadka_roti_rice_combo_1770013924380.png"],
    [20, "Indian Mains", "Yellow Dal Tadka", "Slow-cooked yellow lentils tempered with cumin, garlic and desi ghee. Simple, comforting and full of homely flavour.", 299, 349, "No", "dal_tadka_roti_rice_combo_1770013924380.png"],
    [21, "Indian Mains", "Dal Makhani", "Slow-cooked creamy dal.", 349, 399, "Yes", "dal_tadka_combo_thali_1770012594484.png"],
    [22, "Indian Mains", "Paneer Butter Masala", "Soft paneer cubes cooked in a rich tomato-based gravy with butter and mild spices. A classic hotel favourite that pairs perfectly with rice or rotis.", 399, 449, "Yes", "paneer_butter_masala_combo_bowl_1770012612956.png"],
    [25, "Indian Mains", "Palak Paneer", "Fresh spinach gravy blended with aromatic spices and soft paneer, prepared in traditional hotel-style for a smooth and comforting taste.", 399, 449, "No", "paneer_butter_masala_combo_bowl_1770012612956.png"],
    [23, "Breads", "Butter Roti", "Whole wheat roti.", 45, 60, "No", "breakfast_paratha_combo_1770012647403.png"],
    [24, "Breads", "Aloo Paratha", "Stuffed potato paratha.", 129, 159, "Yes", "breakfast_paratha_combo_1770012647403.png"],
    [30, "Desserts", "Gajar Ka Halwa", "Slow-cooked carrot dessert prepared in traditional style, lightly sweetened and served warm.", 199, 249, "Yes", "gajar_halwa_bowl_1770013800745.png"],
    [31, "Desserts", "Rice Kheer", "Creamy rice pudding gently flavoured with cardamom, prepared fresh for a comforting dessert experience.", 149, 199, "No", "gajar_halwa_bowl_1770013800745.png"],
    [32, "Juices", "Fresh Fruit Juice", "Freshly prepared juice made from seasonal fruits. No artificial flavours, no concentrates.", 149, 199, "No", "fresh_fruit_juices_assorted_1770013780221.png"]
]

for row in menu_data:
    ws_menu.append(row)

# Styling Header
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="EF4F5F", end_color="EF4F5F", fill_type="solid") # Zomato Red
for cell in ws_menu[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

# --- Sheet 2: Daily SOP Checklist ---
ws_sop = wb.create_sheet("Daily SOP Checklist")
ws_sop.append(["ID", "Category", "Task", "Frequency", "Assigned To", "Status (Checked)"])

sop_data = [
    [1, "Ops Control", "Accepting Orders Promptly", "Daily", "Manager", ""],
    [2, "Shift Prep", "Morning Breakfast Prep (7-10 AM)", "Daily", "Chef", ""],
    [3, "Packaging", "Soup: Sealed Container + Outer Cover", "Per Order", "Packer", ""],
    [4, "Packaging", "Rice/Noodles: Rectangular Box", "Per Order", "Packer", ""],
    [5, "Packaging", "Biryani: Wide Cont. + Raita Separate", "Per Order", "Packer", ""],
    [6, "Packaging", "Soup Lids Taped Securely", "Per Order", "Packer", ""],
    [7, "Profit Control", "Portion Spoon Fixed (Cost Control)", "Continuous", "Chef", ""],
    [8, "Profit Control", "Oil Usage Controlled", "Continuous", "Chef", ""],
    [9, "Quality", "Every Order Checked Before Seal", "Per Order", "Manager", ""],
    [10, "Quality", "Raita Always with Biryani", "Per Order", "Packer", ""],
    [11, "Provisions", "Check Flour Stock", "Daily", "Store Keeper", ""],
    [12, "Provisions", "Check Salt Stock", "Daily", "Store Keeper", ""],
    [13, "Provisions", "Check Pepper Stock", "Daily", "Store Keeper", ""],
    [14, "Provisions", "Check Cheese Stock", "Daily", "Store Keeper", ""]
]

for row in sop_data:
    ws_sop.append(row)

# Style SOP Header
for cell in ws_sop[1]:
    cell.font = header_font
    cell.fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid") # Dark Grey
    cell.alignment = Alignment(horizontal="center")


# --- Sheet 3: Sales Register (Template) ---
ws_sales = wb.create_sheet("Sales Register")
ws_sales.append(["Date", "Order ID", "Customer Name", "Items Ordered", "Total Amount (INR)", "Payment Mode", "Status"])
ws_sales.append([datetime.now().strftime("%Y-%m-%d"), "#ORD-001", "Rahul Sharma", "Dal Tadka Combo (x2)", 1598, "UPI", "Delivered"])

# Style Sales Header
for cell in ws_sales[1]:
    cell.font = header_font
    cell.fill = PatternFill(start_color="24963F", end_color="24963F", fill_type="solid") # Veg Green

# --- Sheet 4: Professional Staff Attendance & Timesheet ---
ws_staff = wb.create_sheet("Staff Attendance & Timesheet")
staff_headers = [
    "Date", "Employee ID", "Employee Name", "Role", "Shift Type", 
    "Check-In Time", "Check-Out Time", "Break (Mins)", 
    "Gross Hours", "Net Work Hours", "Overtime (Hrs)", 
    "Status", "Daily Wage (INR)", "Total Pay (INR)"
]
ws_staff.append(staff_headers)

# Employee Roster
employees = [
    ("EMP-001", "Rajesh Kumar", "Head Chef", 1200),
    ("EMP-002", "Sunil Singh", "Sous Chef", 900),
    ("EMP-003", "Amit Verma", "Commi 1", 700),
    ("EMP-004", "Vikram Das", "Commi 2", 600),
    ("EMP-005", "Suresh Yadav", "Kitchen Helper", 500),
    ("EMP-006", "Deepak Koi", "Kitchen Helper", 500),
    ("EMP-007", "Rohan Mehta", "Store Keeper", 800),
    ("EMP-008", "Priya Sharma", "Manager", 1500),
    ("EMP-009", "Karan Johar", "Delivery Partner", 500),
    ("EMP-010", "Arjun Rampal", "Delivery Partner", 500)
]

# Generate data for past 7 days
import random
from datetime import timedelta

current_date = datetime.now()
dates = [(current_date - timedelta(days=i)).strftime("%Y-%m-%d") for i in range(7)]

row_idx = 2
for date in dates:
    for emp_id, emp_name, role, wage in employees:
        # Simulate realistic variations
        status = "Present"
        if random.random() < 0.05: status = "Absent" # 5% chance of absence
        elif random.random() < 0.10: status = "Half Day" # 5% chance of half day

        if status == "Absent":
            row_data = [date, emp_id, emp_name, role, "-", "-", "-", 0, 0, 0, 0, "Absent", wage, 0]
        else:
            # Shift Logic - 4 Shift Patterns to cover 24 hours (mostly active hours)
            shift_choice = random.choice([1, 2, 3, 4])
            
            if shift_choice == 1:
                shift_type = "Shift 1 (Morning)"
                base_in_hour = 7
            elif shift_choice == 2:
                shift_type = "Shift 2 (Day)"
                base_in_hour = 11
            elif shift_choice == 3:
                shift_type = "Shift 3 (Evening)"
                base_in_hour = 14
            else:
                shift_type = "Shift 4 (Night)"
                base_in_hour = 17

            # Check In variations (base time +/- 15 mins)
            in_hour = base_in_hour
            in_min = random.randint(0, 30)
            check_in = f"{in_hour:02d}:{in_min:02d}"
            
            # Check Out variations (Work ~9 hours)
            work_hours = 9 + random.uniform(-0.5, 1.5) # 8.5 to 10.5 hours
            if status == "Half Day": work_hours = 4.5
            
            out_total_min = (in_hour * 60) + in_min + int(work_hours * 60)
            
            # Handle day rollover (24-hour format)
            out_hour_raw = (out_total_min // 60)
            out_hour = out_hour_raw % 24
            out_min = out_total_min % 60
            
            check_out = f"{out_hour:02d}:{out_min:02d}"
            
            # If shift crosses midnight (e.g., Night shift), add Next Day indicator strictly for logic if needed, 
            # but for simple timesheet text, just valid hours is fine.

            
            break_mins = 45 if status == "Present" else 15
            gross_hours = work_hours
            net_hours = gross_hours - (break_mins/60)
            overtime = max(0, net_hours - 9)
            
            daily_pay = wage if status == "Present" else (wage / 2)
            total_pay = daily_pay + (overtime * (wage/9 * 1.5)) # 1.5x OT pay

            row_data = [
                date, emp_id, emp_name, role, shift_type,
                check_in, check_out, break_mins,
                round(gross_hours, 2), round(net_hours, 2), round(overtime, 2),
                status, wage, round(total_pay, 2)
            ]

        ws_staff.append(row_data)
        row_idx += 1

# Style Staff Header
for cell in ws_staff[1]:
    cell.font = header_font
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid") # Blue
    cell.alignment = Alignment(horizontal="center")

# Adjust Column Widths
for ws in [ws_menu, ws_sop, ws_sales, ws_staff]:
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

# Save
file_path = "Sky5_Kitchen_Admin_Kit.xlsx"
wb.save(file_path)
print(f"Excel file created at {file_path}")
