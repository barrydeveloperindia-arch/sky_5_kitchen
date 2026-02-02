from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Create workbook and sheet
wb = Workbook()
ws = wb.active
ws.title = "Combo Sale Menu"

# Styling definitions
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="E23744", end_color="E23744", fill_type="solid") # Zomato Red-ish
center_align = Alignment(horizontal="center", vertical="center")
left_align = Alignment(horizontal="left", vertical="center")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Headers
headers = [
    "Combo Category",
    "Combo Name",
    "Items Included",
    "Sale Price (INR)",
    "Mark as Best Seller"
]

ws.append(headers)

# Apply header styling
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

# Combo data from User Request
combos = [
    ("Veg Comfort Combo", "Dal Tadka Combo", "Dal Tadka + Jeera Rice + 2 Butter Roti", 799, "Yes"),
    ("Veg Comfort Combo", "Rajma Chawal Combo", "Rajma + Steamed Rice", 649, "No"),
    ("Veg Comfort Combo", "Khichdi Comfort Bowl", "Khichdi + Papad", 599, "No"),

    ("Paneer Special Combo", "Paneer Butter Masala Combo", "Paneer Butter Masala + Steamed Rice", 849, "Yes"),
    ("Paneer Special Combo", "Palak Paneer Combo", "Palak Paneer + 2 Butter Roti", 799, "No"),

    ("Chinese Combo", "Chilli Paneer Combo", "Chilli Paneer + Veg Fried Rice", 899, "Yes"),
    ("Chinese Combo", "Noodles Combo", "Veg Hakka Noodles + Spring Roll (2 pcs)", 799, "No"),

    ("Breakfast Combo", "Paratha Breakfast Combo", "Aloo / Paneer Paratha + Curd + Butter", 349, "No"),
    ("Breakfast Combo", "Poha Breakfast Combo", "Poha / Upma + Tea / Coffee", 249, "No"),

    ("Dessert Add-on", "Main + Dessert Add-on", "Any Main Course + Kheer / Halwa", 299, "No")
]

for row_data in combos:
    ws.append(row_data)
    for cell in ws[ws.max_row]:
        cell.alignment = left_align
        cell.border = thin_border

# Adjust column widths
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 45
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 20

# Save file
file_path = "Sky_5_Kitchen_Combo_Sale_Menu.xlsx"
wb.save(file_path)
print(f"Excel file created at: {file_path}")
